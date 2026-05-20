from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from policy_transfer.config import SUPPLIER_CHANNEL, SUPPLIER_CHANNEL_CODE, SUPPLIER_USER_ACCOUNT
from policy_transfer.models import PolicyCase, ProductLine


def export_policy_import(template: Path, output: Path, case: PolicyCase) -> None:
    wb = load_workbook(template)
    product_lookup = _product_lookup(wb)

    policy_ws = wb["policy"]
    rider_ws = wb["rider"]
    _clear_data(policy_ws)
    _clear_data(rider_ws)

    main = next((p for p in case.products if p.kind == "basic"), case.products[0] if case.products else ProductLine("basic"))
    policy_row = _policy_row(policy_ws, case, main, product_lookup)
    _write_row(policy_ws, 2, policy_row)

    rider_row_index = 2
    for rider in [p for p in case.products if p.kind == "rider"]:
        _write_row(rider_ws, rider_row_index, _rider_row(rider_ws, case, rider, product_lookup))
        rider_row_index += 1

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def _clear_data(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def _headers(ws) -> dict[str, int]:
    return {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}


def _write_row(ws, row_index: int, values: dict[str, Any]) -> None:
    headers = _headers(ws)
    if ws.max_row >= 2:
        for cell in ws[2]:
            ws.cell(row=row_index, column=cell.column)._style = copy(cell._style)
    for header, value in values.items():
        col = headers.get(header)
        if col:
            ws.cell(row=row_index, column=col, value=value)


def _policy_row(ws, case: PolicyCase, product: ProductLine, product_lookup: dict[str, tuple[str, str]]) -> dict[str, Any]:
    product_name, product_code = _resolve_product(product, product_lookup)
    sign_date = _date(case.sign_date.value)
    policy_date = _date(case.policy_date.value or case.sign_date.value)
    first_premium = _date(case.first_premium_date.value)
    proposer = case.proposer
    insured = case.insured
    return {
        "保單號碼*": case.policy_no.value or case.proposal_no.value,
        "保單狀態*": case.policy_status.value or "INFORCE",
        "簽單日期": sign_date,
        "遞交日期": _date(case.submission_date.value) or sign_date,
        "內部遞交日期": _date(case.submission_date.value) or sign_date,
        "繕發日期": _date(case.issue_date.value),
        "保單日期*": policy_date,
        "首期保費日": first_premium,
        "下期保費日": _date(case.next_premium_date.value),
        "合約日期*": _date(case.commission_date.value) or policy_date,
        "冷靜期結束日期": _date(case.cooling_off_end_date.value),
        "保單回溯": False,
        "DDA": False,
        "是否轉入": True,
        "保險公司*": "保誠",
        "保險公司碼*": "PRUDENTIAL",
        "產品名稱*": product_name,
        "產品碼*": product_code,
        "付款年期*": product.premium_term.value or "",
        "付款频率*": case.payment_mode.value or "ANNUALLY",
        "每期保费*": product.modal_premium.value or case.total_modal_premium.value,
        "保額*": product.sum_assured.value or "",
        "貨幣*": case.currency.value or "USD",
        "病房等級（醫療險適用）": product.ward_level.value or "",
        "保障地區（醫療險適用）": product.medical_region.value or "",
        "自付額（醫療險適用）": product.deductible.value or "",
        "附加保障（醫療險適用）": product.rider_benefit.value or "",
        "供應渠道*": SUPPLIER_CHANNEL,
        "供應渠道碼*": SUPPLIER_CHANNEL_CODE,
        "用戶賬號*": SUPPLIER_USER_ACCOUNT,
        "業務代表1*": _tr_account(case),
        "申請人類型*": "INDIVIDUAL",
        "投保人是否合資格投資者*": False,
        "投保人中文姓*": proposer.chinese_name.value[:1] if proposer.chinese_name.value else "",
        "投保人中文名*": proposer.chinese_name.value[1:] if proposer.chinese_name.value else "",
        "投保人英文姓*": proposer.english_family_name.value,
        "投保人英文名*": proposer.english_given_name.value,
        "投保人出生日期*": _date(proposer.date_of_birth.value),
        "投保人性別*": proposer.sex.value,
        "投保人國籍*": proposer.nationality.value,
        "投保人聯繫電話區號": proposer.phone_country_code.value,
        "投保人聯繫電話": proposer.phone.value,
        "投保人聯繫郵箱": proposer.email.value,
        "投保人身份證號/出生證*": proposer.id_number.value,
        "投保人通行證號碼": proposer.travel_permit_number.value,
        "投保人戶籍地址": proposer.residential_address.value,
        "投保人居住地址": proposer.residential_address.value,
        "投保人通訊地址": proposer.correspondence_address.value or proposer.residential_address.value,
        "投保人公司名稱": proposer.employer.value,
        "投保人公司行業類型": proposer.business_nature.value,
        "投保人公司地址": proposer.business_address.value,
        "投保人職位": proposer.occupation.value,
        "投保人月薪": case.financial.monthly_income.value,
        "投保人是否願意接受電郵推廣宣傳": True,
        "投保人是否擁有其他國家公民身份（如美國、日本等）": False,
        "與投保人關係*": _relationship_code(case.relationship.value),
        "受保人類型*": "INDIVIDUAL",
        "受保人中文姓*": insured.chinese_name.value[:1] if insured.chinese_name.value else "",
        "受保人中文名*": insured.chinese_name.value[1:] if insured.chinese_name.value else "",
        "受保人英文姓*": insured.english_family_name.value,
        "受保人英文名*": insured.english_given_name.value,
        "受保人出生日期*": _date(insured.date_of_birth.value),
        "受保人性別*": insured.sex.value,
        "受保人國籍*": insured.nationality.value,
        "受保人身份證號/出生證*": insured.id_number.value,
        "受保人戶籍地址*": insured.residential_address.value,
        "受保人居住地址": insured.residential_address.value,
        "受保人通訊地址": insured.correspondence_address.value or insured.residential_address.value,
        "受保人是否願意接受電郵推廣宣傳": True,
        "受保人是否擁有其他國家公民身份（如美國、日本等）": False,
        "受益人1與受保人關係*": "",
        "受益人1受益比例": case.beneficiary_share.value,
        "內部備註": "Generated by local policy transfer tool; review low-confidence fields before C system import.",
    }


def _rider_row(ws, case: PolicyCase, rider: ProductLine, product_lookup: dict[str, tuple[str, str]]) -> dict[str, Any]:
    product_name, product_code = _resolve_product(rider, product_lookup)
    return {
        "保單號碼*": case.policy_no.value or case.proposal_no.value,
        "保險公司*": "保誠",
        "保險公司碼*": "PRUDENTIAL",
        "產品名稱*": product_name,
        "產品碼*": product_code,
        "付款年期*": rider.premium_term.value or "",
        "每期保费*": rider.modal_premium.value or "",
        "保額*": rider.sum_assured.value or "N/A",
        "BPO": False,
        "病房等級（醫療險適用）": rider.ward_level.value or "",
        "保障地區（醫療險適用）": rider.medical_region.value or "",
        "自付額（醫療險適用）": rider.deductible.value or "",
        "附加保障（醫療險適用）": rider.rider_benefit.value or "",
    }


def _product_lookup(wb) -> dict[str, tuple[str, str]]:
    lookup: dict[str, tuple[str, str]] = {}
    ws = wb["product"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        insurer_code = row[3]
        traditional = row[5]
        code = row[7]
        if insurer_code == "PRUDENTIAL" and traditional and code:
            lookup[str(traditional)] = (str(traditional), str(code))
    return lookup


def _resolve_product(product: ProductLine, lookup: dict[str, tuple[str, str]]) -> tuple[str, str]:
    name = str(product.name.value or "")
    for key, value in lookup.items():
        if name and (name in key or key in name):
            return value
    return (name, str(product.code.value or ""))


def _date(value: object) -> str | None:
    if not value:
        return None
    return str(value).replace("-", "/")


def _relationship_code(value: object) -> str:
    text = str(value or "")
    if "Grandparent" in text or "祖父" in text:
        return "GRANDPARENT"
    if "Parent" in text or "父母" in text:
        return "PARENT"
    if "Own" in text or "本人" in text:
        return "OWN"
    return text


def _tr_account(case: PolicyCase) -> str:
    name = str(case.tr_name.value or "").lower().replace(" ", ".")
    return name
