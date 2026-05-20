from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from policy_transfer.exporters import export_bundle
from policy_transfer.exporters.excel_export import _tr_account
from policy_transfer.extractors import ExtractionInput, PrudentialExtractor
from policy_transfer.models import FieldValue, Person, PolicyCase, ProductLine
from policy_transfer.server import _apply_tr_override


import os


ROOT = Path(os.environ.get("POLICY_TRANSFER_SAMPLE_ROOT", "/Users/anthony/Documents/transfer"))
A_FILES = [ROOT / "A" / "Document.pdf", ROOT / "A" / "Document-3.pdf"]
EPOLICY_FILE = ROOT / "A" / "ePolicy_ZENG DONGLING_14260829 (ZENG01896130).PDF"


def main() -> None:
    if not all(path.exists() for path in A_FILES):
        print("Sample A PDFs not found; set POLICY_TRANSFER_SAMPLE_ROOT to run sample-based tests.")
        test_future_extractor_shape()
        print("Portable tests passed.")
        return
    test_prudential_extract()
    test_prudential_epolicy_priority()
    test_export_bundle()
    test_no_tr_account_fallback()
    test_tr_manual_override()
    test_tr_override_empty_keeps_document_value()
    test_future_extractor_shape()
    print("All tests passed.")


def _inputs() -> list[ExtractionInput]:
    return [ExtractionInput(path.name, path.read_bytes()) for path in A_FILES]


def test_prudential_extract() -> None:
    case = PrudentialExtractor().extract(_inputs())
    assert case.proposal_no.value == "000014260829", case.proposal_no
    assert case.proposer.english_family_name.value == "ZENG"
    assert case.proposer.english_given_name.value == "DONGLING"
    assert case.proposer.id_number.value == "440821197001240049"
    assert case.insured.english_family_name.value == "OU"
    assert case.insured.english_given_name.value == "MINGHAO"
    assert case.insured.id_number.value == "Q440627592"
    assert case.currency.value == "USD"
    assert case.payment_mode.value == "ANNUALLY"
    assert float(case.total_modal_premium.value) == 595.81
    assert case.products, "products should be extracted"


def test_prudential_epolicy_priority() -> None:
    if not EPOLICY_FILE.exists():
        return
    case = PrudentialExtractor().extract([ExtractionInput(EPOLICY_FILE.name, EPOLICY_FILE.read_bytes())])
    assert case.policy_no.value == "000014260829"
    assert case.policy_date.value == "2026-03-31"
    assert case.first_premium_date.value == "2026-04-01"
    assert float(case.total_modal_premium.value) == 595.21
    premiums = {product.code.value: product.modal_premium.value for product in case.products}
    assert float(premiums["CIM3"]) == 238.05
    assert float(premiums["MLP"]) == 357.16


def test_export_bundle() -> None:
    case = PrudentialExtractor().extract(_inputs())
    with tempfile.TemporaryDirectory() as tmp:
        files = export_bundle(case, Path(tmp))
        for path in files.values():
            assert path.exists(), path
            assert path.stat().st_size > 0, path
        PdfReader(str(files["client_acknowledgement"]))
        PdfReader(str(files["risk_assessment"]))
        booklet_fields = PdfReader(str(files["client_booklet"])).get_fields() or {}
        assert booklet_fields["Dropdown_fn2"].get("/V") in ("", None)
        assert booklet_fields["Check Box_Product"].get("/V") == "/Off"
        booklet_reader = PdfReader(str(files["client_booklet"]))
        assert b"policy-transfer diagonal skip mark" not in booklet_reader.pages[4]._get_contents_as_bytes()
        assert b"policy-transfer diagonal skip mark" in booklet_reader.pages[5]._get_contents_as_bytes()
        assert b"policy-transfer diagonal skip mark" in booklet_reader.pages[11]._get_contents_as_bytes()
        assert b"policy-transfer diagonal skip mark" not in booklet_reader.pages[12]._get_contents_as_bytes()
        wb = load_workbook(files["policy_import"], data_only=True)
        assert "policy" in wb.sheetnames
        assert wb["policy"]["C2"].value == "000014260829"


def test_no_tr_account_fallback() -> None:
    case = PolicyCase()
    assert _tr_account(case) == ""
    case.tr_name = FieldValue("Jane Advisor", 1)
    assert _tr_account(case) == "jane.advisor"


class _FakeForm:
    def __init__(self, values):
        self.values = values

    def getfirst(self, key):
        return self.values.get(key)


def test_tr_manual_override() -> None:
    case = PolicyCase()
    case.tr_name = FieldValue("OLD TR", 0.8)
    case.tr_license_no = FieldValue("OLD123", 0.8)
    _apply_tr_override(
        case,
        _FakeForm(
            {
                "tr_mode": "manual",
                "manual_tr_name": "NEW TR",
                "manual_tr_license_no": "NEW456",
            }
        ),
    )
    assert case.tr_name.value == "NEW TR"
    assert case.tr_license_no.value == "NEW456"
    assert case.tr_name.confidence == 1.0


def test_tr_override_empty_keeps_document_value() -> None:
    case = PolicyCase()
    case.tr_name = FieldValue("DOCUMENT TR", 0.8)
    case.tr_license_no = FieldValue("DOC123", 0.8)
    _apply_tr_override(case, _FakeForm({"tr_mode": "manual"}))
    assert case.tr_name.value == "DOCUMENT TR"
    assert case.tr_license_no.value == "DOC123"


def test_future_extractor_shape() -> None:
    class MockExtractor:
        company_key = "mock"
        display_name = "Mock Insurer"

        def matches(self, files):
            return True

        def extract(self, files):
            case = PolicyCase()
            case.source_company = FieldValue("MOCK", 1)
            case.proposal_no = FieldValue("MOCK-1", 1)
            case.policy_no = FieldValue("MOCK-1", 1)
            case.currency = FieldValue("USD", 1)
            case.payment_mode = FieldValue("ANNUALLY", 1)
            case.total_modal_premium = FieldValue(1000, 1)
            case.proposer = Person(role="proposer", english_family_name=FieldValue("CHAN", 1), english_given_name=FieldValue("TAI MAN", 1), id_number=FieldValue("A1234567", 1), date_of_birth=FieldValue("1980-01-01", 1), sex=FieldValue("MALE", 1), nationality=FieldValue("China", 1))
            case.insured = Person(role="insured", english_family_name=FieldValue("CHAN", 1), english_given_name=FieldValue("TAI MAN", 1), id_number=FieldValue("A1234567", 1), date_of_birth=FieldValue("1980-01-01", 1), sex=FieldValue("MALE", 1), nationality=FieldValue("China", 1), residential_address=FieldValue("Hong Kong", 1))
            case.products = [ProductLine("basic", name=FieldValue("Mock Product", 1), code=FieldValue("MOCK:P1", 1), premium_term=FieldValue(10, 1), sum_assured=FieldValue(100000, 1), modal_premium=FieldValue(1000, 1))]
            return case

    case = MockExtractor().extract([])
    assert case.products[0].name.value == "Mock Product"
    assert case.proposer.id_number.value == "A1234567"


if __name__ == "__main__":
    main()
